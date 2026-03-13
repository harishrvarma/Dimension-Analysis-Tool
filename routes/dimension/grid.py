from flask import Blueprint, jsonify, render_template, request
from services.dimension import grid


grid_bp = Blueprint("grid_bp", __name__, url_prefix="/grid")


@grid_bp.get("")
@grid_bp.get("/")
def grid_page():
    return render_template("dimension/grid/index.html", active_page="grid")


@grid_bp.get("/api/product-groups")
def api_product_groups():
    """Get all product groups with default selection"""
    groups, default_group_id = grid.get_product_groups()
    return jsonify({"groups": groups, "default_group_id": default_group_id})


@grid_bp.post("/api/options")
def api_options():
    """Get brands and categories based on filters"""
    payload = request.get_json(silent=True) or {}
    group_id = payload.get("group_id")
    brands = payload.get("brands") or []
    categories = payload.get("categories") or []
    types = payload.get("types") or []
    final_status = payload.get("final_status") or []
    skip_status = payload.get("skip_status") or []

    if not group_id:
        return jsonify({
            "ok": False,
            "message": "No product group selected.",
            "brand_options": [],
            "category_options": [],
            "type_options": [],
            "analyzed_status": {}
        })

    brand_options = grid.get_brands_with_counts(group_id, final_status if final_status else None)
    category_options = grid.get_categories_with_counts(group_id, brands if brands else None, final_status if final_status else None)
    type_options = grid.get_types_with_counts(group_id, brands if brands else None, categories if categories else None, final_status if final_status else None)
    analyzed_status = grid.get_analyzed_status(group_id, brands if brands else None, categories if categories else None)

    return jsonify({
        "ok": True,
        "brand_options": brand_options,
        "category_options": category_options,
        "type_options": type_options,
        "analyzed_status": analyzed_status,
        "message": f"Loaded options for group {group_id}"
    })


@grid_bp.post("/api/iteration-filters")
def api_iteration_filters():
    """Get brands and categories from iteration"""
    payload = request.get_json(silent=True) or {}
    iteration_id = payload.get("iteration_id")
    
    if not iteration_id:
        return jsonify({"ok": False, "message": "Iteration ID required"})
    
    filters = grid.get_iteration_filters(iteration_id)
    
    if not filters:
        return jsonify({"ok": False, "message": "Iteration not found"})
    
    return jsonify({"ok": True, "filters": filters})


@grid_bp.post("/api/grid-data")
def api_grid_data():
    """Load grid data based on filters"""
    payload = request.get_json(silent=True) or {}
    group_id = payload.get("group_id")
    brands = payload.get("brands") or []
    categories = payload.get("categories") or []
    types = payload.get("types") or []
    final_status = payload.get("final_status") or []
    skip_status = payload.get("skip_status") or []
    clusters = payload.get("clusters") or []
    iteration_id = payload.get("iteration_id")
    page = payload.get("page", 1)
    per_page = payload.get("per_page", 50)
    sort_column = payload.get("sort_column")
    sort_direction = payload.get("sort_direction", "asc")

    if not group_id:
        return jsonify({"ok": False, "message": "No product group selected.", "data": [], "total": 0})

    data, total = grid.load_grid_data(group_id, brands, categories, types, final_status if final_status else None, skip_status if skip_status else None, clusters if clusters else None, iteration_id, page, per_page, sort_column, sort_direction)

    return jsonify({
        "ok": True,
        "data": data,
        "total": total,
        "message": f"Loaded {len(data)} products"
    })


@grid_bp.post("/api/update-skip-status")
def api_update_skip_status():
    """Update skip status for a single product"""
    payload = request.get_json(silent=True) or {}
    product_id = payload.get("product_id")
    skip_status = payload.get("skip_status")

    if product_id is None:
        return jsonify({"ok": False, "message": "Product ID is required."})

    from models.base.base import SessionLocal
    from repositories.dimension.product_repository import ProductRepository

    db = SessionLocal()
    try:
        repo = ProductRepository(db)
        repo.update_skip_status(product_id, skip_status)
        db.commit()
        return jsonify({"ok": True, "message": "Skip status updated successfully"})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "message": f"Error: {str(e)}"})
    finally:
        db.close()


@grid_bp.post("/api/export-data")
def api_export_data():
    """Export grid data to CSV"""
    payload = request.get_json(silent=True) or {}
    group_id = payload.get("group_id")
    if not group_id:
        return jsonify({"ok": False, "message": "No product group selected."}), 400

    import csv
    from io import StringIO
    from flask import make_response

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Product ID', 'QB Code', 'Brand', 'Category', 'Product Type',
                     'Name', 'Height', 'Width', 'Depth', 'EPS', 'Sample', 'Final Status', 
                     'Total Items', 'Analyzed Items', 'Pending Items', 'Outlier Items',
                     'Cluster Items', 'Cluster Items (%)', 'Cluster', 'Skip Status', 'Final Status History'])

    page = 1
    chunk_size = 5000
    while True:
        data, _ = grid.load_grid_data(group_id, payload.get("brands"), payload.get("categories"),
                                      payload.get("types"), payload.get("final_status"), payload.get("skip_status"),
                                      payload.get("clusters"), payload.get("iteration_id"), page, chunk_size, None, 'asc', skip_count=True)
        if not data:
            break

        for row in data:
            skip_display = 'Yes' if row['skip_status'] == 1 else ('No' if row['skip_status'] == 0 else '-')
            history_text = ''
            if row['iteration_history']:
                history_lines = [f"EPS: {h['eps']}, Sample: {h['sample']}, Status: {h['status']}, {h['date']}"
                               for h in row['iteration_history']]
                history_text = ' | '.join(history_lines)

            writer.writerow([
                row['system_product_id'], row['qb_code'], row['brand'], row['category'], row['product_type'], row['name'],
                row['height'], row['width'], row['depth'], row['eps'], row['sample'], row['final_status'],
                row.get('total_items', 0), row.get('analyzed_items', 0), row.get('pending_items', 0), row.get('outlier_items', 0),
                row.get('cluster_items', 0), f"{row.get('cluster_items_per', 0):.2f}%", row.get('cluster', ''),
                skip_display, history_text
            ])

        if len(data) < chunk_size:
            break
        page += 1

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=grid_export.csv"
    output.headers["Content-type"] = "text/csv"
    return output


@grid_bp.post("/api/export-xls")
def api_export_xls():
    """Export grid data to XLS with red background for outlier dimensions"""
    payload = request.get_json(silent=True) or {}
    group_id = payload.get("group_id")
    if not group_id:
        return jsonify({"ok": False, "message": "No product group selected."}), 400

    from io import BytesIO
    from flask import make_response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        return jsonify({"ok": False, "message": "openpyxl not installed"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Export"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    headers = ['Product ID', 'QB Code', 'Brand', 'Category', 'Product Type',
               'Name', 'Height', 'Width', 'Depth', 'EPS', 'Sample', 'Final Status',
               'Total Items', 'Analyzed Items', 'Pending Items', 'Outlier Items',
               'Cluster Items', 'Cluster Items (%)', 'Cluster', 'Skip Status', 'Final Status History']
    ws.append(headers)

    page = 1
    chunk_size = 5000
    current_row = 2
    cells_to_color = []

    while True:
        data, _ = grid.load_grid_data(group_id, payload.get("brands"), payload.get("categories"),
                                      payload.get("types"), payload.get("final_status"), payload.get("skip_status"),
                                      payload.get("clusters"), payload.get("iteration_id"), page, chunk_size, None, 'asc', skip_count=True)
        if not data:
            break

        for row in data:
            skip_display = 'Yes' if row['skip_status'] == 1 else ('No' if row['skip_status'] == 0 else '-')
            history_text = ''
            if row['iteration_history']:
                history_lines = [f"EPS: {h['eps']}, Sample: {h['sample']}, Status: {h['status']}, {h['date']}"
                               for h in row['iteration_history']]
                history_text = ' | '.join(history_lines)

            ws.append([
                row['system_product_id'], row['qb_code'], row['brand'], row['category'], row['product_type'], row['name'],
                row['height'], row['width'], row['depth'], row['eps'], row['sample'], row['final_status'],
                row.get('total_items', 0), row.get('analyzed_items', 0), row.get('pending_items', 0), row.get('outlier_items', 0),
                row.get('cluster_items', 0), f"{row.get('cluster_items_per', 0):.2f}%", row.get('cluster', ''),
                skip_display, history_text
            ])

            if row['iqr_height_status'] == 0:
                cells_to_color.append((current_row, 7))
            if row['iqr_width_status'] == 0:
                cells_to_color.append((current_row, 8))
            if row['iqr_depth_status'] == 0:
                cells_to_color.append((current_row, 9))

            current_row += 1

        if len(data) < chunk_size:
            break
        page += 1

    for row_num, col_num in cells_to_color:
        ws.cell(row=row_num, column=col_num).fill = red_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=grid_export.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@grid_bp.post("/api/save-skip-status")
def api_save_skip_status():
    """Save skip status for selected products"""
    payload = request.get_json(silent=True) or {}
    skip_items = payload.get("skip_items") or []

    if not skip_items:
        return jsonify({"ok": False, "message": "No items to save."})

    from models.base.base import SessionLocal
    from models.dimension.product import Product

    db = SessionLocal()
    try:
        product_ids = [item["product_id"] for item in skip_items]
        products = db.query(Product).filter(Product.product_id.in_(product_ids)).all()

        skip_map = {item["product_id"]: item["skip_status"] for item in skip_items}

        for product in products:
            if product.product_id in skip_map:
                product.skip_status = skip_map[product.product_id]

        db.commit()
        return jsonify({"ok": True, "message": f"Saved skip status for {len(skip_items)} products"})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "message": f"Error: {str(e)}"})
    finally:
        db.close()
